from torch.utils.tensorboard import SummaryWriter
from pix2Topix import pix2pixG_256, pix2pixD_256, pix2pixG_128, pix2pixD_128
import argparse
from mydatasets import CreateDatasets
from split_data import split_data
import os
from torch.utils.data.dataloader import DataLoader
import torch
import torch.optim as optim
import torch.nn as nn
from utils import train_one_epoch, val
import openpyxl as xl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import glob


def train(opt):
    batch = opt.batch
    data_path = opt.dataPath
    # train_data = opt.train_data
    # test_data = opt.test_data
    print_every = opt.every
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    epochs = opt.epoch
    img_size = opt.imgsize
    numworker = opt.numworker
    result_path = opt.result_path
    save_path = opt.savePath

    if not os.path.exists(save_path):
        os.mkdir(save_path)

    # 加载数据集
    train_imglist, val_imglist = split_data(data_path)
    # 直接使用划分好的train和test
    # train_imglist = []
    # val_imglist= []
    # train_img = glob.glob(train_data + '/*.png')
    # val_img = glob.glob(test_data + '/*.png')
    # for img in train_img:
    #     train_imglist.append(img)
    # for img in val_img:
    #     val_imglist.append(img)
    train_datasets = CreateDatasets(train_imglist, img_size)
    val_datasets = CreateDatasets(val_imglist, img_size)

    train_loader = DataLoader(dataset=train_datasets, batch_size=batch, shuffle=True, num_workers=numworker,
                              drop_last=True)
    val_loader = DataLoader(dataset=val_datasets, batch_size=batch, shuffle=True, num_workers=numworker,
                            drop_last=True)

    # 实例化网络
    pix_G = pix2pixG_256().to(device)
    pix_D = pix2pixD_256().to(device)
    #pix_G = pix2pixG_128().to(device)
    #pix_D = pix2pixD_128().to(device)

    # 定义优化器和损失函数
    optim_G = optim.Adam(pix_G.parameters(), lr=0.0002, betas=(0.5, 0.999))
    optim_D = optim.Adam(pix_D.parameters(), lr=0.0002, betas=(0.5, 0.999))
    loss = nn.MSELoss()
    l1_loss = nn.L1Loss()
    start_epoch = 0

    # 加载预训练权重
    if opt.weight != '':
        ckpt = torch.load(opt.weight)
        pix_G.load_state_dict(ckpt['G_model'], strict=False)
        pix_D.load_state_dict(ckpt['D_model'], strict=False)
        start_epoch = ckpt['epoch'] + 1
    writer_path = 'train_logs_' + str(img_size) + '_' + str(batch) + '_' + str(epochs)
    writer = SummaryWriter(writer_path)

    train_G = []
    train_D = []
    val_G = []
    val_D = []
    # 开始训练
    for epoch in range(start_epoch, epochs):
        loss_mG, loss_mD = train_one_epoch(G=pix_G, D=pix_D, train_loader=train_loader,
                                           optim_G=optim_G, optim_D=optim_D, writer=writer, loss=loss, device=device,
                                           plot_every=print_every, epoch=epoch, l1_loss=l1_loss)

        writer.add_scalars(main_tag='train', tag_scalar_dict={
            'loss_G': loss_mG,
            'loss_D': loss_mD
        }, global_step=epoch)

        train_G.append(loss_mG.item())
        train_D.append(loss_mD.item())
        # train_G.append(loss_mG)
        # train_D.append(loss_mD)

        # 验证集
        loss_mG, loss_mD = val(G=pix_G, D=pix_D, val_loader=val_loader, loss=loss, l1_loss=l1_loss,
                               device=device, epoch=epoch, writer=writer, plot_every=print_every, path=result_path)

        writer.add_scalars(main_tag='val', tag_scalar_dict={
            'loss_G': loss_mG,
            'loss_D': loss_mD
        }, global_step=epoch)

        val_G.append(loss_mG.item())
        val_D.append(loss_mD.item())
        # val_G.append(loss_mG)
        # val_D.append(loss_mD)
    # 保存模型
    save_name = save_path + '/' + str(img_size) + '_' + str(batch) + '_' + str(epochs) + '_lambda=500.pth'
    torch.save({
        'G_model': pix_G.state_dict(),
        'D_model': pix_D.state_dict(),
        'epoch': epoch
    }, save_name)

    return train_G, train_D, val_G, val_D

def draw_loss(train_G, train_D, val_G, val_D):
    epoch = len(train_G)
    fig = plt.figure(figsize=(10, 8))
    x = range(1, epoch + 1)
    ax1 = fig.add_subplot(2, 1, 1)
    ax1.set_title('train_loss')
    ax1.plot(x, train_G, label="train_G_loss", ls="-", marker="o")
    ax1.plot(x, train_D, label="train_D_loss", ls="-", marker="o")
    ax2 = fig.add_subplot(2, 1, 2)
    ax2.set_title('val_loss')
    ax2.plot(x, val_G, label="val_G_loss", ls="-", marker="o")
    ax2.plot(x, val_D, label="val_D_loss", ls="-", marker="o")
    save_pic = opt.result_path + '/loss.png'
    save_xlsx = opt.result_path + '/loss.xlsx'
    plt.savefig(save_pic)
    plt.show()
    wb = xl.Workbook()
    ws1 = wb.create_sheet('train')
    ws2 = wb.create_sheet('val')
    for i in range(epoch):
        ws1.cell(row=i+1, column=1, value=train_G[i])
        ws1.cell(i+1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row=i+1, column=2, value=train_D[i])
        ws1.cell(i+1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=i+1, column=1, value=val_G[i])
        ws2.cell(i+1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=i+1, column=2, value=val_D[i])
        ws2.cell(i+1, 2).alignment = Alignment(horizontal='center', vertical='center')
    wb.save(save_xlsx)


def cfg():
    parse = argparse.ArgumentParser()
    parse.add_argument('--batch', type=int, default=16)
    parse.add_argument('--epoch', type=int, default=100)
    parse.add_argument('--imgsize', type=int, default=256)
    parse.add_argument('--dataPath', type=str, default='./data', help='data root path')
    # parse.add_argument('--train_data', type=str, default='./data', help='train data root path')
    # parse.add_argument('--test_data', type=str, default='./test', help='val data root path')
    parse.add_argument('--result_path', type=str, default='./results_256_16_100_lambda=50', help='results root path')
    parse.add_argument('--weight', type=str, default='', help='load pre train weight') # 'weights/pix2pix_256.pth'
    parse.add_argument('--savePath', type=str, default='./my_weights', help='weight save path')
    parse.add_argument('--numworker', type=int, default=4)
    parse.add_argument('--every', type=int, default=1, help='plot train result every * iters')
    opt = parse.parse_args()
    return opt


if __name__ == '__main__':
    opt = cfg()
    print(opt)
    train_G, train_D, val_G, val_D = train(opt)
    draw_loss(train_G, train_D, val_G, val_D)
